import os
from openpyxl.utils import get_column_letter
import pandas as pd
import warnings

# 打开当前文件夹下的排单表.xlsx
current_directory = os.getcwd()
source_file_name = '资料表.xlsx'
source_file_path = os.path.join(current_directory, source_file_name)


# 生成财务看的总表，该表不同之处在于复制检测项目列后粘贴在最后一列，从显示数量转变为显示具体价格
print("正在生成财务-排单明细表")

# 消除pandas的提示
warnings.filterwarnings('ignore')

finance_excel_path = os.path.join(os.getcwd(), "财务-排单明细表.xlsx")
# shutil.copy(source_file_path, finance_excel_path)

df = pd.read_excel(source_file_path, sheet_name='排单明细')
df = df.sort_values(by=['集团编号', '序号'], ignore_index=True)

p1 = pd.read_excel(source_file_path, sheet_name='价目表1')
p2 = pd.read_excel(source_file_path, sheet_name='价目表2')

price_project_names = list(p1.columns)

# 记录集团编号
group_number = df['集团编号']

# 删除最后四列，该四列不是检测项目，2025年4月9日新增3列
df = df.iloc[:, :-7]

# 检测项目的表头
headers = list(df.columns[21:])


# 选择第22列到最后一列的数据
columns_to_copy = df.columns[21:]  # 从第22列开始到最后一列

# 获取DataFrame的列数
before_copy_last_column = df.shape[1]

# 复制选择的列并追加到最后一列后面
df = pd.concat([df, df[columns_to_copy]], axis=1)

after_copy_last_columns = df.shape[1]


column_dict = {index: title for index, title in enumerate(headers, start=before_copy_last_column+1)}

for index, title in column_dict.items():
    try:
        title_index = price_project_names.index(title)
    except:
        pass
    # 追加excel公式用来计算检测项目费用
    for i in range(0, df.shape[0]):
        # 如果检测项目存在于价目表1
        if title in price_project_names:
            
            amount = str(df.iat[i, index-1])
            # 遇到空白的单元格时，设置为0
            if amount is None or amount == 'nan':
                amount = 0
            # 这里的df.iat[i, index-1]中的i不+1和后面的i需要加2，区别在于前面的i默认忽略的表头，相当于默认加了1，后面的i没有这个前提。
            df.iat[i, index-1] = f"=VLOOKUP(K{i+2},价目表1!$A$2:${get_column_letter(p1.shape[1])}${p1.shape[0]+1},{title_index+1},FALSE)*{amount}"
        # 如果检测项目不存在于价目表1
        else:
            amount = str(df.iat[i, index-1])
            # 遇到空白的单元格时，设置为0
            if amount is None or amount == 'nan':
                amount = 0
            df.iat[i, index-1] = f"=VLOOKUP({get_column_letter(index)}1,价目表2!$A$2:$B${p2.shape[0]+1},2,FALSE)*{amount}"

# 添加 "数量小计" 列，只包含表头，不包含数据
df['数量小计'] = None
# 添加 "金额小计" 列，只包含表头，不包含数据
df['金额小计'] = None
# 添加"集团编号"列
df['集团编号'] = None

for i in range(0, df.shape[0]):
    # 计算数量小计
    df.iat[i, df.shape[1]-3] = f"=SUM({get_column_letter(22)}{i+2}:{get_column_letter(before_copy_last_column)}{i+2})"
    # 计算金额小计
    df.iat[i, df.shape[1]-2] = f"=SUM({get_column_letter(before_copy_last_column+1)}{i+2}:{get_column_letter(after_copy_last_columns)}{i+2})"
    # 插入集团编号
    df.iat[i, df.shape[1]-1] = group_number[i]

# 使用 ExcelWriter 保存多个 DataFrame 到同一份 Excel 文件
print("生成完成，正在写入文件中...", end="")
with pd.ExcelWriter(finance_excel_path) as writer:
    df.to_excel(writer, sheet_name='使用明细', index=False)
    p1.to_excel(writer, sheet_name='价目表1', index=False)
    p2.to_excel(writer, sheet_name='价目表2', index=False)
input("完成。")

# -*- coding: utf-8 -*-
# @Time    : 2023-8-29
# @Update  : 2023-10-8
# @Author  : simonoct14@outlook.com
# @Purposes: 将资料表的排单明细中的数据(手动导入)根据集团编号分类后生成以集团、账期区分的明细文件(使用明细sheet、明细表sheet)。然后将所有与集团对账单(对账单sheet，该文件由别的系统生成)合并，最终形成所有在排单明细中的集团的对账单(对账单sheet、使用明细sheet、明细表sheet)。
# @Comment  : 使用时的版本，numpy=1.26.0 openpyxl=3.1.2 pandas=2.1.1 xlwings=0.30.12 pyinstaller=6.0.0 Python=3.12.0。
import openpyxl
import os
import shutil
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
import pandas as pd
import calendar
import xlwings as xw
import glob


class CategorizeExcel:

    def __init__(self, source_file_path, template, out_file_path, date):
        self.source_file_path = source_file_path
        self.template = template
        self.out_file_path = out_file_path
        """记录生成后的文件的路径"""
        self.fileslist = []
        """保存已经打开的文件对象, 详见execute中的self.output_wb = openpyxl.load_workbook(file)"""
        self.output_wb = None
        """用来记录排序后的对象, 详见sort_data"""
        self.detail_pd_sheet = None
        """用来记录明细表中, 第三行最后一列是多少列"""
        self.template_details_sheet_max_column = 0
        """用来记录使用明细表B列中的检测项目名字"""
        self.template_usage_sheet_column_b = []
        """用来记录gen_formula生成的E列公式"""
        self.usage_sheet_column_e_formual = []
        """用来记录生成的明细表最后一行是第几行, insert_formula会用到, 用来使用明细表的C列公式生成"""
        self.output_detail_last_row = 0
        """用来记录价目表的价格"""
        self.price_dict = {}
        self.price2_dict = {}
        """用来记录价目表1的检测项目名称"""
        self.price_project_names = []
        """打开数据源工作簿"""
        self.source_wb = openpyxl.load_workbook(
            self.source_file_path, read_only=True)
        self.detail_sheet = self.source_wb['排单明细']
        """检查模板中的使用明细的检测项目和明细表的检测项目是否完全一致"""
        self.check_template(template)
        """用来记录协议编号"""
        self.detial_H_4 = ''
        """用来记录协议单位"""
        self.detial_I_4 = ''
        """生成D、E列公式, 因为D、E列公式是固定的。还有生成价目表的字典用于搜寻对应(协议、检测项目)的价格后写入对应的生成文件"""
        self.gen_formula()
        self.date = date
        """对数据表中的数据进行排序后再进行生成文件, 避免存在数据不按集团名顺序排列的情况"""
        self.sort_data(self.source_file_path)
        self.execute()
        """生成财务看的总表，该表不同之处在于检测项目列从显示数量转变为显示具体价格"""
        self.finance_details()

    def sort_data(self, source_file_path):
        print("正在对数据进行排序...")
        """读取Excel文件中的排单明细数据"""
        df = pd.read_excel(source_file_path, sheet_name='排单明细')

        """进行order by操作, 不自动添加索引"""
        self.detail_pd_sheet = df.sort_values(
            by=['集团编号', '序号'], ignore_index=True)
        print("完成。")

    def check_template(self, template_file):
        """检查资料表的排单明细表表头、模板的明细表表头的检查项和使用明细的B列检测项目是否按顺序一一对应(区分大小写)"""
        print("检查资料表的排单明细表、模板里使用明细和明细表的检测项目是否一一对应...")
        template_wb = openpyxl.load_workbook(template_file, read_only=True)
        template_usage_sheet = template_wb['使用明细']
        template_details_sheet = template_wb['明细表']
        detail_sheet = self.detail_sheet
        template_detials_header = []
        source_data_header = []

        """获取最后一列列号"""
        max_column = template_details_sheet.max_column
        """明细表项目列表。19代表S列, max_column - 3指最后4列不计算在内。"""
        for i in range(19, max_column - 3):
            cell_value = template_details_sheet.cell(row=3, column=i).value
            template_detials_header.append(cell_value)

        template_usage_sheet_column_b = []

        for row in range(4, template_usage_sheet.max_row+1):

            cell = template_usage_sheet.cell(row, 2)

            template_usage_sheet_column_b.append(cell.value)

        """排单明细表项目列表。19代表S列, max_column - 3指最后4列不计算在内。"""
        for i in range(19, detail_sheet.max_column - 3):
            cell_value = detail_sheet.cell(row=1, column=i).value
            source_data_header.append(cell_value)

        """模板的检测项目的顺序和内容, 使用明细表和明细表是一致的"""
        if template_detials_header != template_usage_sheet_column_b:
            try:
                for i in range(len(template_usage_sheet_column_b)):
                    if template_usage_sheet_column_b[i] != template_detials_header[i]:
                        self.source_wb.close()
                        template_wb.close()
                        input(
                            f'请检查「使用明细表」和「明细表」的“{template_usage_sheet_column_b[i]}”处左右的项目有无问题! 修正后重新执行! ')
                        exit()
            except Exception as e:
                self.source_wb.close()
                template_wb.close()
                print("模板检查出了错误，请检查模板中使用明细表和明细表的检测项目是否为一一对应的关系！")
                print("下面是错误信息：")
                print(e)
                input("修正后重新执行!")
                exit()
        """资料表的检测项目的顺序和内容, 排单明细表和明细表是一致的"""
        if template_detials_header != source_data_header:
            try:
                for i in range(len(source_data_header)):
                    if source_data_header[i] != template_detials_header[i]:
                        self.source_wb.close()
                        template_wb.close()
                        input(
                            f'请检查「排单明细表」和「明细表」的“{source_data_header[i]}”处左右的项目有无问题! 修正后重新执行! ')
                        exit()
            except Exception as e:
                self.source_wb.close()
                template_wb.close()
                print("资料表和模板检查出了错误，请检查资料表的排单明细和模板明细表的检测项目是否为一一对应的关系！")
                print("下面是错误信息：")
                print(e)
                input("修正后重新执行!")
                exit()
        template_wb.close()
        print('通过。')
        self.template_usage_sheet_column_b = template_usage_sheet_column_b
        self.template_details_sheet_max_column = template_details_sheet.max_column

    def data_row(self, row):
        """返回当前数据行的明细内容"""
        return list(self.detail_pd_sheet.loc[row].values)

    def output_excel(self, file_path):
        """创建文件后打开, 然后记录文件路径用来后续处理"""
        shutil.copy(self.template, file_path)
        self.output_wb = openpyxl.load_workbook(file_path)
        self.fileslist.append(file_path)

    def detial_sum(self):
        """在明细表内底部追加合计一栏和修改A2日期"""

        ws = self.output_wb['明细表']

        """修改A2日期"""
        year = self.date[:4]
        month = self.date[4:]
        days_in_month = calendar.monthrange(int(year), int(month))[1]
        formatted_prompt = f"{year}年{month}月01日~{year}年{month}月{days_in_month}日测试明细"
        ws.cell(row=2, column=1).value = formatted_prompt

        """获取最后一行行号"""
        last_row = ws.max_row
        max_column = ws.max_column
        """计算公式的起始列(R列)"""
        start_col = ord('R') - ord('A') + 1
        """生成公式列表"""
        formula = ['合计']
        for i in range(19, max_column - 1):
            sum = 0
            for row in range(4, last_row + 1):
                """计算合计"""
                number = ws.cell(row=row, column=i).value
                if number is None:
                    number = 0
                sum = number + sum
            formula.append(sum)
        """从第R列开始追加公式"""
        for i, f in enumerate(formula):
            ws.cell(row=last_row+1, column=start_col+i).value = f
            ws.cell(row=last_row+1, column=start_col+i).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            ws.cell(row=last_row+1, column=start_col+i).font = Font(size=10, name='宋体', bold=True)
            border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
            ws.cell(row=last_row+1, column=start_col+i).border = border
            if f == 0:
                ws.column_dimensions.group(get_column_letter(start_col+i), get_column_letter(start_col+i), hidden=True)
        ws.cell(row=1, column=1).font = Font(size=16, name='宋体', bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='bottom')
        ws.column_dimensions.group(get_column_letter(ws.max_column-1), get_column_letter(ws.max_column), hidden=True)
        """打印区域"""
        ws.print_area = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
        self.detial_H_4 = ws.cell(row=4, column=8).value
        self.detial_I_4 = ws.cell(row=4, column=9).value


        """记录明细表的最后一行行数"""
        self.output_detail_last_row = last_row + 1

    def gen_formula(self):
        """D列价格获取公式生成"""
        pirce_sheet = self.source_wb["价目表1"]
        """去掉头尾"""
        price_header_values = pirce_sheet[1][1:-1]
        """将价目表1的检测项目用列表记录下来"""
        price_project_names = []
        for name in price_header_values:
            price_project_names.append(name.value)
        self.price_project_names = price_project_names
        """使用pandas将价目表1、价目表2内的数据转换成类似数据表的形式, 其中协议编号、检测项目为主键"""
        price_sheet = pd.read_excel(self.source_file_path, sheet_name="价目表1")
        price2_sheet = pd.read_excel(self.source_file_path, sheet_name="价目表2")
        """将"column1"列设置为索引"""
        price_sheet.set_index("协议编号", inplace=True)
        price2_sheet.set_index("检测项目", inplace=True)
        """将DataFrame转换为字典"""
        self.price_dict = price_sheet.to_dict(orient="index")
        self.price2_dict = price2_sheet.to_dict(orient="index")

        """使用明细的数据列从第4行开始, 生成E列的公式"""
        count = 4
        for i in self.template_usage_sheet_column_b:
            self.usage_sheet_column_e_formual.append(f'=C{count}*D{count}')
            count += 1

    def insert_formula(self):
        output_usage_sheet = self.output_wb['使用明细']
        output_details_sheet = self.output_wb['明细表']
        agreement_number = self.output_wb['明细表']['H4'].value
        output_details_sheet.column_dimensions.group('H', 'I', hidden=True)
        output_details_sheet.column_dimensions.group('O', 'O', hidden=True)
        output_usage_sheet.cell(row=1, column=2).value = self.detial_I_4
        output_usage_sheet.cell(row=2, column=2).value = self.detial_H_4
        usage_sheet_data_start_row = 4
        """A、C、D、E列数目公式插入"""
        row = usage_sheet_data_start_row
        count = 0
        """range(19, self.template_details_sheet_max_column - 3)的长度和self.price_project_names是一样的(check_template), 这里用前者是因为C列公式需要指定明细表的合计行。"""
        for i in range(19, self.template_details_sheet_max_column - 3):
            c_formula = output_details_sheet.cell(row=self.output_detail_last_row, column=i).value
            output_usage_sheet.cell(row=row, column=3).value = c_formula
            """判断当前行的检测项目是在价目表1还是价目表2, 然后在D列插入对应的价格"""
            try:
                if output_usage_sheet.cell(row=row, column=2).value in self.price_project_names:
                    row_data_list = list(
                        self.price_dict.get(agreement_number).values())
                    index = self.price_project_names.index(
                        output_usage_sheet.cell(row=row, column=2).value)
                    d_formula = row_data_list[index]
                else:
                    row_data_list = list(self.price2_dict.get(
                        output_usage_sheet.cell(row=row, column=2).value).values())
                    d_formula = row_data_list[0]
            except:
                self.finished()
                input("请检查协议编号或检测项目是否存在于价目表1或检测项目是否存在于价目表2! 确认无误后请关闭该窗口, 重新运行该程序! ")
                exit()
            output_usage_sheet.cell(row=row, column=4).value = d_formula
            """插入E列公式"""
            e_formula = self.usage_sheet_column_e_formual[count]
            output_usage_sheet.cell(row=row, column=5).value = e_formula

            """根据输入日期修改A列日期"""
            year = self.date[:4]
            month = self.date[4:]
            formatted_date = year + "年" + str(int(month)) + "月"
            output_usage_sheet.cell(row=row, column=1).value = formatted_date

            if c_formula == 0:
                output_usage_sheet.row_dimensions.group(row, row, hidden=True)

            row += 1
            count += 1

        """在使用明细追加合计行"""
        output_usage_sheet.cell(row=output_usage_sheet.max_row+1,
                                column=3).value = '="共 "&SUM(C4:C{0})&" 款"'.format(output_usage_sheet.max_row)
        output_usage_sheet.cell(row=output_usage_sheet.max_row, column=5).value = '="合计 "&SUM(E4:E{0})&" 元"'.format(
            output_usage_sheet.max_row-1)
        """设置文字格式"""
        alignment = Alignment(horizontal='center')
        font = Font(color="FF0000", bold=True, size=14)
        for col in output_usage_sheet.iter_cols(min_row=output_usage_sheet.max_row, max_row=output_usage_sheet.max_row):
            for cell in col:
                cell.alignment = alignment
                cell.font = font
        
        """打印区域"""
        output_usage_sheet.print_area = f'A1:{get_column_letter(output_usage_sheet.max_column)}{output_usage_sheet.max_row}'

    def finished(self):

        if self.output_wb:
            self.output_wb.close()
        self.source_wb.close()

    def execute(self):
        print("正在处理数据中...")
        """往明细表填入数据"""
        row = 1
        data_list = []
        """插入第一条数据用来比较"""
        data_list.append(self.data_row(0))
        """当row计数小于detail_pd_sheet的最大行数时, 则一直循环, self.detail_pd_sheet为数据表排序后的对象, 不含表头"""
        while row < self.detail_pd_sheet.shape[0]:
            """将每一行的数据(列表)追加到data_list中"""
            data = self.data_row(row)
            data_list.append(data)
            """如果现在的索引8与上一次索引8不一样, 意味着集团名不一样, 则表明上一个集团已经完成记录, 需要写入到文件中。"""
            if data_list[len(data_list)-2][8] != data_list[len(data_list)-1][8]:
                """构建文件名"""
                group_number = data_list[len(data_list)-2][self.template_details_sheet_max_column-2]
                agreement_number = data_list[len(data_list)-2][7]
                group_name = data_list[len(data_list)-2][8]
                file_name = f"{group_number}-{agreement_number}-{group_name}[{self.date}].xlsx"
                file_path = os.path.join(self.out_file_path, file_name)
                """创建文件"""
                self.output_excel(file_path)
                """将本次循环记录的数据跳出来(不属于同一个集团), 用last_data记录"""
                last_data = data_list.pop()
                """将数据一条条写入到文件中"""
                number = 1
                for data in data_list:
                    ws = self.output_wb['明细表']
                    ws.append(data)
                    ws.cell(row=number+3, column=1).value = number
                    for column in range(1, len(data)+1):
                        ws.cell(row=number+3, column=column).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                        ws.cell(row=number+3, column=column).font = Font(size=10, name='宋体')
                        border = Border(left=Side(border_style='thin', color='000000'),
                                        right=Side(border_style='thin', color='000000'),
                                        top=Side(border_style='thin', color='000000'),
                                        bottom=Side(border_style='thin', color='000000'))
                        ws.cell(row=number+3, column=column).border = border
                    number += 1
                """清空data_list, 然后将本次记录的数据重新插入到data_list中"""
                number = 1
                data_list.clear()
                data_list.append(last_data)
                """保存文件"""
                self.output_wb.save(file_path)
                self.output_wb.close()

            row += 1

        """循环后, 最后一个列表需要插入, 无论最后一条数据是否与前一次循环的相同, 都需要执行"""
        group_number = data_list[len(data_list)-1][self.template_details_sheet_max_column-2]
        agreement_number = data_list[len(data_list)-1][7]
        group_name = data_list[len(data_list)-1][8]
        file_name = f"{group_number}-{agreement_number}-{group_name}[{self.date}].xlsx"
        file_path = os.path.join(self.out_file_path, file_name)
        self.output_excel(file_path)
        print("完成。")
        number = 1
        for data in data_list:
            ws = self.output_wb['明细表']
            ws.append(data)
            ws.cell(row=number+3, column=1).value = number
            for column in range(1, len(data)+1):
                ws.cell(row=number+3, column=column).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                ws.cell(row=number+3, column=column).font = Font(size=10, name='宋体')
                border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
                ws.cell(row=number+3, column=column).border = border
            number += 1
        number = 1
        self.output_wb.save(file_path)
        self.output_wb.close()

        print(f"共生成了{len(self.fileslist)}份文件。")
        self.output_wb.close()

        print("正在往文件插入公式...")
        for file in self.fileslist:
            """往明细表最后添加合计"""
            self.output_wb = openpyxl.load_workbook(file)
            self.detial_sum()
            """往使用明细表插入的C、D、E列插入公式"""
            self.insert_formula()

            """修改sheet name"""
            detials_sheet = self.output_wb['使用明细']
            usage_sheet = self.output_wb['明细表']

            detials_sheet.title = f'{self.date}使用明细'
            usage_sheet.title = f'{self.date}明细表'

            self.output_wb.save(file)
            self.output_wb.close()
        print("完成。")
    
    def finance_details(self):
        print("正在生成财务-排单明细表.xlsx")
        finance_excel_path = os.path.join(os.getcwd(), "财务-排单明细表.xlsx")
        shutil.copy(source_file_path, finance_excel_path)

        wb = openpyxl.load_workbook(finance_excel_path)
        ws = wb['排单明细']

        # 获取表头行
        header_row = ws[1]

        # 删除最后四列，该四列不是检测项目
        ws.delete_cols(ws.max_column - 3, 4)

        # 检验项目开始的列数
        start_column = 19
        # 记录复制列前的最大列数
        before_copy_max_column = ws.max_column

        # 复制从第19列到最后一列的数据
        data_to_copy = []
        for col in ws.iter_cols(min_col=19, values_only=True):
            data_to_copy.append(col)

        # 追加复制的数据到最后一列，并设置字体和对齐方式
        for col_data in data_to_copy:
            start = 1
            max_column = ws.max_column + 1
            
            for value in col_data:
                ws.cell(row=start, column=max_column, value=value)
                if start == 1:
                    ws.cell(row=start, column=max_column).font = Font(size=10, name='宋体', bold=True)
                else:
                    ws.cell(row=start, column=max_column).font = Font(size=10, name='宋体')
                ws.cell(row=start, column=max_column).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                start += 1

        # 提取需要插入公式的表头列
        selected_columns = header_row[start_column-1:]
        selected_columns = selected_columns[:-4]

        # 使用enumerate将列索引与对应的列标题组成字典
        column_dict = {index: title for index, title in enumerate(selected_columns, start=before_copy_max_column+1)}

        for index, title in column_dict.items():

            # 尝试获取表头索引，self.price_project_names是由价目表1的表头组成的列表
            try:
                title_index = self.price_project_names.index(title.value)
            except:
                pass
            # 追加excel公式用来计算检测项目费用
            for i in range(2, ws.max_row+1):
                # 如果检测项目存在于价目表1
                if title.value in self.price_project_names:
                    
                    amount = ws.cell(row=i, column=index).value
                    # 遇到空白的单元格时，设置为0
                    if amount is None:
                        amount = 0
                    ws.cell(row=i, column=index).value = f"=VLOOKUP(H{i},价目表1!$A$2:${get_column_letter(wb['价目表1'].max_column)}${wb['价目表1'].max_row},{title_index+2},FALSE)*{amount}"
                # 如果检测项目不存在于价目表1
                else:
                    amount = ws.cell(row=i, column=index).value
                    # 遇到空白的单元格时，设置为0
                    if amount is None:
                        amount = 0
                    ws.cell(row=i, column=index).value = f"=VLOOKUP({get_column_letter(index)}1,价目表2!$A$2:$B${wb['价目表2'].max_row},2,FALSE)*{amount}"

        # 记录追加公式前最大列数
        before_sum_max_column = ws.max_column

        # 在列表最后添加数量合计列
        ws.cell(row=1, column=ws.max_column+1).value = "数量小计"
        ws.cell(row=1, column=ws.max_column).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        ws.cell(row=1, column=ws.max_column).font = Font(size=10, name='宋体', bold=True)
        start_sum_column = get_column_letter(start_column)
        end_sum_column = get_column_letter(before_copy_max_column)
        for row in range(2, ws.max_row+1):
            ws.cell(row=row, column=ws.max_column).value = f'=SUM({start_sum_column}{row}:{end_sum_column}{row})'
            ws.cell(row=row, column=ws.max_column).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            ws.cell(row=row, column=ws.max_column).font = Font(size=10, name='宋体')

        # 在列表最后添加金额合计列
        ws.cell(row=1, column=ws.max_column+1).value = "金额小计"
        ws.cell(row=1, column=ws.max_column).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        ws.cell(row=1, column=ws.max_column).font = Font(size=10, name='宋体', bold=True)
        start_currency_sum_column = get_column_letter(before_copy_max_column+1)
        end_currency_sum_column = get_column_letter(before_sum_max_column)
        for row in range(2, ws.max_row+1):
            ws.cell(row=row, column=ws.max_column).value = f'=SUM({start_currency_sum_column}{row}:{end_currency_sum_column}{row})'
            ws.cell(row=row, column=ws.max_column).number_format = '#,##0.00'
            ws.cell(row=row, column=ws.max_column).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            ws.cell(row=row, column=ws.max_column).font = Font(size=10, name='宋体')

        wb.save(finance_excel_path)
        wb.close()
        print("完成。")


# 打开当前文件夹下的排单表.xlsx
current_directory = os.getcwd()

# 构建文件路径

while True:
    date = input("请输入账期(格式为yyyymm): ")
    if len(date) == 6 and date.isdigit():
        break
    else:
        print("日期格式错误, 请重新输入。")
year = date[:4]
month = date[4:]

source_file_name = '资料表.xlsx'
output_folder = f'{year}年{month}月'
merge_folder = f'{date}对账单-完整版'
template_file_name = '模板.xlsx'
statements_folder = '简易系统对账单'


source_file_path = os.path.join(current_directory, source_file_name)

output_folder_path = os.path.join(current_directory, output_folder)
merge_folder_path = os.path.join(current_directory, merge_folder)
statements_folder_path = os.path.join(current_directory, statements_folder)


template = os.path.join(current_directory, template_file_name)

if os.path.exists(output_folder_path):
    print(f"请检查{output_folder}目录有无文件, 如有的话请先将文件夹内的文件移动到别处或删除, 否则程序将退出。")
    input("按回车键继续...")
    try:
        os.rmdir(output_folder_path)
    except:
        exit()

if os.path.exists(merge_folder_path):
    print(f"请检查{merge_folder}目录有无文件, 如有的话请先将文件夹内的文件移动到别处或删除, 否则程序将退出。")
    input("按回车键继续...")
    try:
        os.rmdir(merge_folder_path)
    except:
        exit()

if not os.path.exists(statements_folder_path):
    input("请检查“简易系统对账单”文件夹是否已经准备就绪, 准备好后按回车继续...")

if len(os.listdir(statements_folder_path)) == 0:
    input("“简易系统对账单”内无文件，请添加后重新执行程序！")
    exit()

# Create a new folder
os.makedirs(output_folder_path)
os.makedirs(merge_folder_path)

try:
    CategorizeExcel(source_file_path=source_file_path,
                    template=template, out_file_path=output_folder_path, date=date)
except KeyboardInterrupt:
    print(f"你已手动结束程序, 如需重新执行, 请先删除“{output_folder}”、“{merge_folder}”文件夹")
    input("按回车或直接关闭程序...")
    exit()
except Exception as e:
    print("程序出现了错误, 下面是异常信息: ")
    print(e)
    print(f"尝试删除“{output_folder}”、“{merge_folder}”文件夹后重新运行。")
    input("按回车或直接关闭程序...")
    exit()

# 使用glob模块来匹配文件夹内的所有.xlsx文件
out_files = glob.glob(os.path.join(output_folder_path, '*.xlsx'))
statements_files = glob.glob(os.path.join(statements_folder_path, '*.xlsx'))
# 提取文件名
output_files_name = [os.path.basename(file) for file in out_files]
output_files_dict = {}
for file in output_files_name:
    group_num = file.split('-')[0]
    output_files_dict[group_num] = file
statements_files_name = [os.path.basename(file) for file in statements_files]
statements_files_dict = {}
try:
    for file in statements_files_name:
        group_num = file.split('-')[2]
        statements_files_dict[group_num] = file
except IndexError:
    print(f"“{statements_folder}”内有文件存在命名问题，命名要求请查看说明书！")

count = 0
print("正在合并文件中...")

for output_file_group_num, output_file_name in output_files_dict.items():

    statements_file_name = None
    for statements_group_num in statements_files_dict:
        if output_file_group_num == statements_group_num:
            statements_file_name = statements_files_dict[statements_group_num]
            break
        else:
            statements_file_name = None
    if statements_file_name is not None:
        # 打开文件1和文件2
        file1 = os.path.join(statements_folder_path, statements_file_name)
        file2 = os.path.join(output_folder_path, output_file_name)
        # visible用来设置程序是否可见，True表示可见（默认），Flase不可见。
        # add_book用来设置是否自动创建工作簿，True表示自动创建（默认），False不创建。
        # 不过这里给不给参数都可以，只是visible不给的话会一直看到excel程序闪烁。
        try:
            app = xw.App(visible=False, add_book=False)
            wb1 = app.books.open(file1)
            wb2 = app.books.open(file2)

            # 获取文件2的工作表名称
            sheet_names_file2 = [sheet.name for sheet in wb2.sheets]

            # 遍历文件2的工作表，将它们复制到文件1的工作簿中，并追加到文件1的最后
            print(f"正在合并{statements_file_name}文件...", end="")
            for sheet_name in sheet_names_file2:
                sheet2 = wb2.sheets[sheet_name]
                sheet2.api.Copy(After=wb1.sheets[-1].api)

            # 保存修改后的文件1
            wb1.save(f'{os.path.join(merge_folder_path, statements_file_name)}')
            wb1.close()
            wb2.close()
        except:
            print(f"合并出现了错误！请删除“{statements_folder}”文件夹后重新运行合并程序。")
            input("按回车键或直接关闭程序...")
            exit()
        print("完成。")
        count += 1
if count == 0:
    print("注意：本次执行并没有合并任何文件！")
    print(f"①请检查“{statements_folder}”内是否存有文件。 ")
    print(f"②“{statements_folder}”内文件的格式是否为“XX-XX-集团编号-XX.xlsx”。其中“集团编号”是本程序合并的依据! 请确保开头格式为“XX-XX-集团编号-”。")
    print("XX可以为除“-”外的任意字符, 包括空格。")
    print(f"③检查“{output_folder}”内是否有文件。")
    print(f"④确认。“{output_folder}”内公司存在于“{statements_folder}”")
    print(f"⑤检查“{output_folder}”内文件名格式是否为“集团编号-协议号-公司名[日期]”。")
else:
    print("完成。")
    print(f"合并了{count}份文件。")
input("程序执行完毕, 关闭窗口或按回车结束。")
